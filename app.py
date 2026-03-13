import csv
import io
import os
import sqlite3
from datetime import datetime

from flask import (
    Flask,
    g,
    jsonify,
    render_template,
    request,
    send_file,
)

# ── 配置 ──────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(__file__), "repos.db")
CSV_PATH = os.path.join(os.path.dirname(__file__), "top_1000_os_rules.csv")
HOST = "0.0.0.0"
PORT = 5000
DEBUG = True

app = Flask(__name__)


# ── 数据库连接 ────────────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# ── 初始化数据库 & 导入 CSV ──────────────────────────────────────────
def init_db():
    """首次运行时自动建表并导入 CSV 数据。"""
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("""
        CREATE TABLE IF NOT EXISTS repos (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            repo_name       TEXT NOT NULL,
            organization    TEXT,
            language        TEXT,
            stars           INTEGER DEFAULT 0,
            forks           INTEGER DEFAULT 0,
            open_issues     INTEGER DEFAULT 0,
            age_days        INTEGER DEFAULT 0,
            has_coc         INTEGER DEFAULT 0,
            has_contributing INTEGER DEFAULT 0,
            has_workflows   INTEGER DEFAULT 0,
            has_readme      INTEGER DEFAULT 0,
            description     TEXT,
            created_at      TEXT,
            updated_at      TEXT
        )
    """)

    row_count = db.execute("SELECT COUNT(*) FROM repos").fetchone()[0]
    if row_count == 0 and os.path.exists(CSV_PATH):
        now = datetime.now().isoformat()
        with open(CSV_PATH, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = []
            for row in reader:
                rows.append((
                    row.get("repo_name", ""),
                    row.get("organization", ""),
                    row.get("language", ""),
                    int(row.get("stars", 0) or 0),
                    int(row.get("forks", 0) or 0),
                    int(row.get("open_issues", 0) or 0),
                    int(row.get("age_days", 0) or 0),
                    1 if row.get("has_coc", "").strip().lower() == "true" else 0,
                    1 if row.get("has_contributing", "").strip().lower() == "true" else 0,
                    1 if row.get("has_workflows", "").strip().lower() == "true" else 0,
                    1 if row.get("has_readme", "").strip().lower() == "true" else 0,
                    row.get("description", ""),
                    now,
                    now,
                ))
            db.executemany("""
                INSERT INTO repos (
                    repo_name, organization, language, stars, forks,
                    open_issues, age_days, has_coc, has_contributing,
                    has_workflows, has_readme, description, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, rows)
        db.commit()
        print(f"[OK] imported {len(rows)} rows")
    db.close()


# ── 辅助函数 ─────────────────────────────────────────────────────────
def row_to_dict(row):
    return dict(row)


BOOL_FIELDS = {"has_coc", "has_contributing", "has_workflows", "has_readme"}


def parse_bool(val):
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, str):
        return 1 if val.lower() in ("true", "1", "yes") else 0
    return int(bool(val))


def build_filter_conditions(args):
    """从请求参数中构建通用的 WHERE 条件（搜索 + 语言 + 高级筛选）。"""
    conditions = []
    params = []

    search = args.get("search", "").strip()
    language = args.get("language", "").strip()

    if search:
        conditions.append("(repo_name LIKE ? OR description LIKE ? OR organization LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like])
    if language:
        conditions.append("language = ?")
        params.append(language)

    stars_min = args.get("stars_min", type=int)
    stars_max = args.get("stars_max", type=int)
    forks_min = args.get("forks_min", type=int)
    forks_max = args.get("forks_max", type=int)
    age_min = args.get("age_min", type=float)
    age_max = args.get("age_max", type=float)

    if stars_min is not None:
        conditions.append("stars >= ?")
        params.append(stars_min)
    if stars_max is not None:
        conditions.append("stars <= ?")
        params.append(stars_max)
    if forks_min is not None:
        conditions.append("forks >= ?")
        params.append(forks_min)
    if forks_max is not None:
        conditions.append("forks <= ?")
        params.append(forks_max)
    if age_min is not None:
        conditions.append("age_days >= ?")
        params.append(int(age_min * 365))
    if age_max is not None:
        conditions.append("age_days <= ?")
        params.append(int(age_max * 365))

    if args.get("req_readme"):
        conditions.append("has_readme = 1")
    if args.get("req_coc"):
        conditions.append("has_coc = 1")
    if args.get("req_contributing"):
        conditions.append("has_contributing = 1")
    if args.get("req_workflows"):
        conditions.append("has_workflows = 1")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    return where, params


# ── 页面路由 ─────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


# ── API：获取所有语言列表（用于筛选下拉） ───────────────────────────
@app.route("/api/languages")
def api_languages():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT language FROM repos WHERE language != '' ORDER BY language"
    ).fetchall()
    return jsonify([r["language"] for r in rows])


# ── API：列表 & 搜索 ─────────────────────────────────────────────────
@app.route("/api/repos")
def api_list():
    db = get_db()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    sort_by = request.args.get("sort_by", "stars")
    order = request.args.get("order", "desc").upper()

    allowed_sort = {
        "repo_name", "organization", "language", "stars",
        "forks", "open_issues", "age_days", "id",
    }
    if sort_by not in allowed_sort:
        sort_by = "stars"
    if order not in ("ASC", "DESC"):
        order = "DESC"

    where, params = build_filter_conditions(request.args)

    total = db.execute(f"SELECT COUNT(*) FROM repos {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = db.execute(
        f"SELECT * FROM repos {where} ORDER BY {sort_by} {order} LIMIT ? OFFSET ?",
        params + [per_page, offset],
    ).fetchall()

    return jsonify({
        "data": [row_to_dict(r) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
    })


# ── API：获取单条 ─────────────────────────────────────────────────────
@app.route("/api/repos/<int:repo_id>")
def api_get(repo_id):
    db = get_db()
    row = db.execute("SELECT * FROM repos WHERE id = ?", (repo_id,)).fetchone()
    if row is None:
        return jsonify({"error": "未找到该记录"}), 404
    return jsonify(row_to_dict(row))


# ── API：新增 ─────────────────────────────────────────────────────────
@app.route("/api/repos", methods=["POST"])
def api_create():
    db = get_db()
    data = request.get_json(force=True)
    now = datetime.now().isoformat()

    for bf in BOOL_FIELDS:
        if bf in data:
            data[bf] = parse_bool(data[bf])

    db.execute("""
        INSERT INTO repos (
            repo_name, organization, language, stars, forks,
            open_issues, age_days, has_coc, has_contributing,
            has_workflows, has_readme, description, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get("repo_name", ""),
        data.get("organization", ""),
        data.get("language", ""),
        int(data.get("stars", 0) or 0),
        int(data.get("forks", 0) or 0),
        int(data.get("open_issues", 0) or 0),
        int(data.get("age_days", 0) or 0),
        data.get("has_coc", 0),
        data.get("has_contributing", 0),
        data.get("has_workflows", 0),
        data.get("has_readme", 0),
        data.get("description", ""),
        now, now,
    ))
    db.commit()
    return jsonify({"message": "创建成功", "id": db.execute("SELECT last_insert_rowid()").fetchone()[0]}), 201


# ── API：更新 ─────────────────────────────────────────────────────────
@app.route("/api/repos/<int:repo_id>", methods=["PUT"])
def api_update(repo_id):
    db = get_db()
    data = request.get_json(force=True)
    now = datetime.now().isoformat()

    existing = db.execute("SELECT * FROM repos WHERE id = ?", (repo_id,)).fetchone()
    if existing is None:
        return jsonify({"error": "未找到该记录"}), 404

    for bf in BOOL_FIELDS:
        if bf in data:
            data[bf] = parse_bool(data[bf])

    db.execute("""
        UPDATE repos SET
            repo_name = ?, organization = ?, language = ?, stars = ?, forks = ?,
            open_issues = ?, age_days = ?, has_coc = ?, has_contributing = ?,
            has_workflows = ?, has_readme = ?, description = ?, updated_at = ?
        WHERE id = ?
    """, (
        data.get("repo_name", existing["repo_name"]),
        data.get("organization", existing["organization"]),
        data.get("language", existing["language"]),
        int(data.get("stars", existing["stars"]) or 0),
        int(data.get("forks", existing["forks"]) or 0),
        int(data.get("open_issues", existing["open_issues"]) or 0),
        int(data.get("age_days", existing["age_days"]) or 0),
        data.get("has_coc", existing["has_coc"]),
        data.get("has_contributing", existing["has_contributing"]),
        data.get("has_workflows", existing["has_workflows"]),
        data.get("has_readme", existing["has_readme"]),
        data.get("description", existing["description"]),
        now,
        repo_id,
    ))
    db.commit()
    return jsonify({"message": "更新成功"})


# ── API：删除 ─────────────────────────────────────────────────────────
@app.route("/api/repos/<int:repo_id>", methods=["DELETE"])
def api_delete(repo_id):
    db = get_db()
    existing = db.execute("SELECT id FROM repos WHERE id = ?", (repo_id,)).fetchone()
    if existing is None:
        return jsonify({"error": "未找到该记录"}), 404
    db.execute("DELETE FROM repos WHERE id = ?", (repo_id,))
    db.commit()
    return jsonify({"message": "删除成功"})


# ── API：统计数据（图表用） ──────────────────────────────────────────
@app.route("/api/stats")
def api_stats():
    db = get_db()

    lang_dist = db.execute("""
        SELECT language, COUNT(*) as count, SUM(stars) as total_stars
        FROM repos WHERE language != ''
        GROUP BY language ORDER BY count DESC LIMIT 15
    """).fetchall()

    stars_ranges = db.execute("""
        SELECT
            CASE
                WHEN stars >= 200000 THEN '200k+'
                WHEN stars >= 100000 THEN '100k-200k'
                WHEN stars >= 50000  THEN '50k-100k'
                WHEN stars >= 20000  THEN '20k-50k'
                ELSE '<20k'
            END as range_label,
            COUNT(*) as count
        FROM repos
        GROUP BY range_label
        ORDER BY MIN(stars) DESC
    """).fetchall()

    top_10 = db.execute(
        "SELECT repo_name, stars FROM repos ORDER BY stars DESC LIMIT 10"
    ).fetchall()

    summary = db.execute("""
        SELECT
            COUNT(*) as total,
            SUM(stars) as total_stars,
            ROUND(AVG(stars)) as avg_stars,
            MAX(stars) as max_stars,
            COUNT(DISTINCT language) as lang_count
        FROM repos
    """).fetchone()

    return jsonify({
        "language_distribution": [row_to_dict(r) for r in lang_dist],
        "stars_ranges": [row_to_dict(r) for r in stars_ranges],
        "top_10": [row_to_dict(r) for r in top_10],
        "summary": row_to_dict(summary),
    })


# ── API：导出 ─────────────────────────────────────────────────────────
@app.route("/api/export/<fmt>")
def api_export(fmt):
    db = get_db()
    where, params = build_filter_conditions(request.args)
    rows = db.execute(
        f"SELECT * FROM repos {where} ORDER BY stars DESC", params
    ).fetchall()

    columns = [
        "id", "repo_name", "organization", "language", "stars", "forks",
        "open_issues", "age_days", "has_coc", "has_contributing",
        "has_workflows", "has_readme", "description",
    ]

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        for r in rows:
            writer.writerow([r[c] for c in columns])
        mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        return send_file(mem, mimetype="text/csv", as_attachment=True,
                         download_name="repos_export.csv")

    elif fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Repos"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(rows, 2):
            for col_idx, c in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=r[c])

        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(col_name) + 4)

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="repos_export.xlsx")

    return jsonify({"error": "不支持的格式，请使用 csv 或 excel"}), 400


# ── 启动 ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    print(f"[START] http://localhost:{PORT}")
    app.run(host=HOST, port=PORT, debug=DEBUG)
